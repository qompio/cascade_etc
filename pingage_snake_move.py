# ----------------------------------------------------------------- #
# Title: Snake Pin Data
# Desc: Program takes column of data and prints it in rows back and forth

# ------------------------------------------------------------------ #
# imports
from openpyxl import load_workbook

# user input path, column, start row
# inpath = str(input("Input file path: ")).strip()
# savepath = str(input("Input save path: ")).strip()
# print()
# column = str(input("Input column number (alphabet --> numeric (E=5): ")).upper().strip()
# start_row = int(input("Input start row: "))

# open work book
path = r"F:\pingagerawdata\0010.xlsx"
wb = load_workbook(path)
# wb = load_workbook('r' + inpath)
ws = wb.active

# create constants
data = []
letters = ["I", "J", "K"]
revletter = letters[::-1]
z = 9


# take data from row and col info and make it into a list
for row in ws.iter_rows(min_row=9, max_col=5, min_col=5):
    for cell in row:
        data.append(cell.value)

# break the list into chunks of 3
lstsize = 3
chunks = [data[i * lstsize:(i + 1) * lstsize] for i in range((len(data) + lstsize - 1))]

# write the snake to excel
for chunk in chunks:
    z += 1
    if z % 2 == 0:
        for i, letter in enumerate(letters):
            try:
                ws[letter + str(z)] = chunk[i]
            except:
                pass

    elif z % 2 != 0:
        for k, letter in enumerate(revletter):
            try:
                ws[letter + str(z)] = chunk[k]
            except:
                pass
    else:
        print("how")

wb.save(r"F:\pingagerawdata\0010_edited.xlsx")
# wb.save('r' + savepath)

