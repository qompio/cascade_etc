from openpyxl import load_workbook

# open work book
# path =  path to workbook
wb = load_workbook(path)
# wb = load_workbook('r' + inpath)
ws = wb.active

range = [*range(28, 247)]

for row in range[::-1]:
    ws.insert_rows(row, 2)

# save edited workbook
# savepath = path to save
# wb.save(savepath)